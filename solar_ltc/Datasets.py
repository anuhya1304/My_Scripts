import streamlit as st
import pandas as pd
import plotly.express as px
from io import BytesIO
from utilsfol.apply_adaption import AdaptionMain
from utilsfol.correlation_calc import CorrelationCalc
from utilsfol.rmse_calc import RMSECalc
from utilsfol.mbe_calc import MBECalc
from ReadData import readdata
from utils import DataHandler, ColumnMapping
import plotly.graph_objects as go
from utilsfol.All_correlation_calc import CorrelationCalcAll
from utilsfol.All_mbe_calc import AllMBECalc
from utilsfol.All_rmse_calc import AllRMSECalc
from openpyxl import load_workbook
import os


st.set_page_config(
    page_title="Energy Index App",
    page_icon="🧊",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://www.extremelycoolapp.com/help',
        'Report a bug': "https://www.extremelycoolapp.com/bug",
        'About': "# This is a header. This is an *extremely* cool app!"
    }
)


@st.cache_data
def cached_load_data(uploaded_file):

    df, data_type, data_dict = readdata(uploaded_file)

    columnmapping = ColumnMapping(df)
    col_mapped_dict = columnmapping.column_flag_mapping()

    datahandler = DataHandler(df, col_mapped_dict)
    if data_type == 'Measured':
        df = datahandler.df_format()
        df = datahandler.interpolate_data()
        df_1hr = datahandler.resample_data()
    if data_type == 'Satellite':
        df_1hr = datahandler.resample_sat_data()

    return df, data_type, data_dict, col_mapped_dict, df_1hr


# @st.cache_data
def cached_applyadap(satellit_df_adap_in, resample_measured_df_hourly_in, mea_ghi_col, sat_ghi_col, data_type):

    common_df = pd.DataFrame()
    satellit_df_adap = satellit_df_adap_in.copy()
    resample_measured_df_hourly = resample_measured_df_hourly_in.copy()

    satellit_df_adap.dropna(subset=['GHI_bins'], inplace=True)
    resample_measured_df_hourly.dropna(subset=['GHI_bins'], inplace=True)

    common_df = pd.merge(resample_measured_df_hourly, satellit_df_adap,
                         left_index=True, right_index=True, how='inner')

    common_df.dropna(subset=[mea_ghi_col, sat_ghi_col], inplace=True)

    resample_measured_df_hourly.reset_index(inplace=True)

    # Apply adaptation methods
    adaptionmain = AdaptionMain(
        satellit_df_adap, resample_measured_df_hourly, common_df, mea_ghi_col, sat_ghi_col, data_type)

    reg_result_adap1, satellit_df_adap1 = adaptionmain.applyadaption_met1()
    reg_result_adap2, satellit_df_adap2 = adaptionmain.applyadaption_met2()

    # Dynamically create the adapted column names based on the data_type
    adapted_col1 = f'{data_type.lower()}_adapted_1'
    adapted_col2 = f'{data_type.lower()}_adapted_2'

    # Print adapted columns for debugging
    print(
        f'Adapted column 1: {adapted_col1}, Adapted column 2: {adapted_col2}')

    # Merging with adapted columns
    common_df = pd.merge(common_df, satellit_df_adap1[[adapted_col1]],
                         left_index=True, right_index=True, how='inner')
    common_df = pd.merge(common_df, satellit_df_adap2[[adapted_col2]],
                         left_index=True, right_index=True, how='inner')

    # Return the updated common_df
    return common_df


# Create tabs
datasettab, adaptiontab = st.tabs(
    ["📝 Dataset", '📊 AdaptionTab'])


timezone_path = './static/time_zone.csv'

timezone_df = pd.read_csv(timezone_path, encoding='ISO-8859-1')
unique_timezones = timezone_df['UTC'].unique()


# Content for Tab 1
with datasettab:
    # Upload multiple files
    uploaded_files = st.file_uploader(
        "Add Datasets", type=['csv', 'txt', 'xlsx'], accept_multiple_files=True)

    files_dict = {}
    data_dicts_list = []

    # Save the uploaded files in session state
    if uploaded_files is not None:
        st.session_state.uploaded_files = uploaded_files

    if st.session_state.uploaded_files is not None:

        cols = st.columns(1 if len(uploaded_files) ==
                          0 else len(uploaded_files))

        for idx, file in enumerate(uploaded_files):
            data_dict1 = {}
            filename = file.name.split('.')[0]
            df, data_type, data_dict, col_mapped_dict_tmp, df_1hr = cached_load_data(
                file)
            if data_type == "Satellite":
                df_1hr = df.copy()
            else:
                col_mapped_dict = col_mapped_dict_tmp
            files_dict[filename] = {
                'df': df, 'data_type': data_type, 'resampled_data': df_1hr}
            data_dict1['File Name'] = filename
            data_dict1['Data Type'] = data_type
            data_dict1.update(data_dict)
            data_dict1['Data Starts Date'] = df.index.min()
            data_dict1['Data Ends Date'] = df.index.max()
            # time_zone = []

            data_dict_df = pd.DataFrame(
                list(data_dict1.items()), columns=['Key', 'Value'])

            files_dict[filename]['data_dict_df'] = data_dict_df

            with cols[idx]:
                st.markdown(
                    f"<h1 style='font-size: 25px;'>{filename}</h1>",
                    unsafe_allow_html=True
                )
                # st.write(data_dict_df)
                st.dataframe(data_dict_df, hide_index=True,
                             use_container_width=True)
                selected_timezone = st.selectbox(
                    f"Please Select the required Time Zone for {filename}",
                    options=unique_timezones
                )
        st.session_state.files_dict = files_dict

    else:
        st.write("Please Upload the file.")


# Adaption_Tab
with adaptiontab:
    if "calculation_button" not in st.session_state:
        st.session_state.calculation_button = None

    adap_datasettab, adap_graphtab, down_csvtab = st.tabs(
        ["📝 Adaption", "📈 Graph", "💾 CSV Download"])

    with adap_datasettab:
        adap_col1, adap_col2 = st.columns(2)

        # Get the Measured and Satellite data types
        measured_files = [
            filename for filename, value in files_dict.items() if value['data_type'] == 'Measured']
        satellite_files = [
            filename for filename, value in files_dict.items() if value['data_type'] == 'Satellite']

        adap_col_sel1, adap_col_sel2 = st.columns(2)
        with adap_col_sel1:
            adap_data_type = st.selectbox(
                'Select Data Type', ['GHI', 'DNI', 'Temp', 'All'], index=3)  # keeoing 'all' as default

        with adap_col1:
            # Display 'Measured' data type
            Measured_tab = st.selectbox(
                'Select the Measured Data:',
                measured_files if measured_files else [
                    "No Measured Data Available"]
            )

        with adap_col2:
            # Display 'Satellite' data type
            Satellite_tab = st.selectbox(
                'Select Satellite Data:',
                satellite_files if satellite_files else [
                    "No Satellite Data Available"]
            )
        if Measured_tab != "No Measured Data Available":
            adap_data_columns_meas = []

            if adap_data_type in ["GHI", "All"]:
                adap_data_columns_meas += [
                    x for x in col_mapped_dict.keys() if x.startswith('GHI')]

            if adap_data_type in ["DNI", "All"]:
                adap_data_columns_meas += [
                    x for x in col_mapped_dict.keys() if x.startswith('DNI')]

            if adap_data_type in ["Temp", "All"]:
                adap_data_columns_meas += [
                    x for x in col_mapped_dict.keys() if x.startswith('Temp')]
        if Satellite_tab != "No Satellite Data Available":
            adap_data_columns_sat = []

            if adap_data_type in ["GHI", "All"]:
                adap_data_columns_sat += [x for x in files_dict[Satellite_tab]
                                          ['df'].columns if x.startswith('GHI')]

            if adap_data_type in ["DNI", "All"]:
                adap_data_columns_sat += [x for x in files_dict[Satellite_tab]
                                          ['df'].columns if x.startswith('DNI')]

            if adap_data_type in ["Temp", "All"]:
                adap_data_columns_sat += [x for x in files_dict[Satellite_tab]
                                          ['df'].columns if x.startswith('Temp')]

        adap_col3, adap_col4 = st.columns(2)

        with adap_col3:
            if Measured_tab != "No Measured Data Available":
                mea_ghi_col = st.selectbox(
                    'Please Select the Measured Data Columns',
                    adap_data_columns_meas
                )
            else:
                st.write("No Measured data selected.")

        with adap_col4:
            if Satellite_tab != "No Satellite Data Available":
                sat_ghi_col = st.selectbox(
                    'Please Select the Satellite Data Columns',
                    adap_data_columns_sat
                )
            else:
                st.write("No Satellite data selected.")

        if Measured_tab == "No Measured Data Available" or Satellite_tab == "No Satellite Data Available":
            st.session_state.calculation_button = None

        if Measured_tab != "No Measured Data Available" and Satellite_tab != "No Satellite Data Available":
            calculation_button = st.button('Apply Adaption')
            # Save the uploaded files in session state
            if calculation_button:
                st.session_state.calculation_button = calculation_button

        if st.session_state.calculation_button:
            combined_corr = {}
            combined_rmse = {}
            combined_mbe = {}

            if adap_data_type == "All":

                data_dict_temp = {}
                # Collect columns for GHI, DNI, Temp into one list
                combined_columns_meas = []
                combined_columns_sat = []

                for dtype in ['GHI', 'DNI', 'Temp']:  # ,

                    combined_columns_meas += [
                        col for col in col_mapped_dict.keys() if col.startswith(dtype)]
                    combined_columns_sat += [
                        col for col in files_dict[Satellite_tab]['df'].columns if col.startswith(dtype)]

                    if dtype == "GHI":
                        mea_adap_col = [
                            x for x in col_mapped_dict.keys() if x.startswith('GHI')][0]
                        # print(len(files_dict[Satellite_tab]['resampled_data']), len(
                        #     files_dict[Measured_tab]['resampled_data']))
                        sat_adap_col = 'GHI'
                        # Apply for GHI
                        common_df_temp = cached_applyadap(
                            files_dict[Satellite_tab]['resampled_data'],
                            files_dict[Measured_tab]['resampled_data'],
                            mea_adap_col,
                            sat_adap_col,
                            data_type='GHI'
                        )
                        data_dict_temp[dtype] = common_df_temp

                    elif dtype == "DNI":
                        mea_adap_col = [
                            x for x in col_mapped_dict.keys() if x.startswith('DNI')][0]

                        sat_adap_col = 'DNI'

                        common_df_temp = cached_applyadap(
                            files_dict[Satellite_tab]['resampled_data'],
                            files_dict[Measured_tab]['resampled_data'],
                            mea_adap_col,
                            sat_adap_col,
                            data_type='DNI'
                        )

                        data_dict_temp[dtype] = common_df_temp
                        # st.write(data_dict_temp)

                        if data_dict_temp:
                            common_df = None
                            for dtype, df in data_dict_temp.items():
                                adapted_columns = [
                                    col for col in df.columns if 'adapted' in col]
                                adapted_df = df[adapted_columns]

                                if common_df is None:
                                    common_df = df.copy()

                                    # common_df = df[[
                                    #     col for col in df.columns if 'adapted' in col]]
                                else:
                                    # Merge adapted columns with common_df, ensuring all columns are retained
                                    merged = pd.merge(common_df, adapted_df,
                                                      left_index=True, right_index=True, how='inner')

                                    # Drop duplicate columns if any (keeping the first occurrence)
                                    merged = merged.loc[:, ~
                                                        merged.columns.duplicated()]

                                    common_df = merged
                        # st.write('merged df is:', common_df)
                        print(
                            "Columns in common_df after merging adapted data:", common_df.columns)

                        if common_df is not None:
                            st.session_state.common_df = common_df
                            # =============================================================================
                            # Apply correlation calculations on the combined common_df
                            # =============================================================================

                            mea_ghi_col = [
                                x for x in col_mapped_dict.keys() if x.startswith('GHI')][0]

                            mea_dni_col = [
                                x for x in col_mapped_dict.keys() if x.startswith('DNI')][0]

                            adap_table_tabletype = st.selectbox(
                                'Select the Data Type:',
                                ['Correlation', 'RMSE', 'MBE']
                            )

                            if adap_table_tabletype == 'Correlation':

                                # Instantiate CorrelationCalcAll with the common_df and necessary columns
                                correlation_calculator = CorrelationCalcAll(
                                    common_df, mea_ghi_col, mea_dni_col)

                                # Get the results for overall, hourly, and monthly correlations
                                correlation_results_overall, correlation_results_hourly, correlation_results_monthly = correlation_calculator.cal_corr()

                                # correlation_results_monthly, correlation_results_hourly, correlation_results_overall = correlation_calculator.cal_corr()
                                st.header('Overall Correlation')
                                st.write(correlation_results_overall)
                                st.header('Hourly Correlation')
                                st.write(correlation_results_hourly)
                                st.header('Monthly Correlation')
                                st.write(correlation_results_monthly)

                            if adap_table_tabletype == 'MBE':
                                all_mbecalc = AllMBECalc(
                                    common_df, mea_ghi_col, mea_dni_col)
                                mbe_results_overall, mbe_results_hourly, mbe_results_monthly, \
                                    mbe_results_overall_per, mbe_results_hourly_per, mbe_results_monthly_per = all_mbecalc.calc_mbe()
                                st.header('Overall MBE Absolute and %')
                                adap_tab_c1, adap_tab_c2 = st.columns(2)
                                with adap_tab_c1:
                                    st.write(mbe_results_overall)
                                with adap_tab_c2:
                                    st.write(mbe_results_overall_per)

                                st.header('Monthly MBE Absolute and %')
                                adap_tab_c3, adap_tab_c4 = st.columns(2)
                                with adap_tab_c3:
                                    st.write(mbe_results_monthly)
                                with adap_tab_c4:
                                    st.write(mbe_results_monthly_per)

                                st.header('Hourly MBE Absolute and %')
                                adap_tab_c5, adap_tab_c6 = st.columns(2)
                                with adap_tab_c5:
                                    st.write(mbe_results_hourly)
                                with adap_tab_c6:
                                    st.write(mbe_results_hourly_per)

                            if adap_table_tabletype == 'RMSE':
                                all_rmsecalc = AllRMSECalc(
                                    common_df, mea_ghi_col, mea_dni_col)
                                rmse_results_overall, rmse_results_hourly, rmse_results_monthly, \
                                    rmse_results_overall_per, rmse_results_hourly_per, rmse_results_monthly_per = all_rmsecalc.calc_rmse()
                                st.header('Overall RMSE Absolute and %')
                                adap_tab_c1, adap_tab_c2 = st.columns(2)
                                with adap_tab_c1:
                                    st.write(rmse_results_overall)
                                with adap_tab_c2:
                                    st.write(rmse_results_overall_per)

                                st.header('Monthly RMSE Absolute and %')
                                adap_tab_c3, adap_tab_c4 = st.columns(2)
                                with adap_tab_c3:
                                    st.write(rmse_results_monthly)
                                with adap_tab_c4:
                                    st.write(rmse_results_monthly_per)

                                st.header('Hourly RMSE Absolute and %')
                                adap_tab_c5, adap_tab_c6 = st.columns(2)
                                with adap_tab_c5:
                                    st.write(rmse_results_hourly)
                                with adap_tab_c6:
                                    st.write(rmse_results_hourly_per)

            else:
                if not adap_data_type == "All":
                    common_df = cached_applyadap(
                        files_dict[Satellite_tab]["resampled_data"], files_dict[Measured_tab]["resampled_data"], mea_ghi_col, sat_ghi_col, data_type=adap_data_type)
                    print('the commondf columns are :', common_df.columns)

                    correlationcalc = CorrelationCalc(common_df, mea_ghi_col)
                    rmsecalc = RMSECalc(common_df, mea_ghi_col)
                    mbecalc = MBECalc(common_df, mea_ghi_col)

                    correlation_results_monthly, correlation_results_hourly, correlation_results_overall = correlationcalc.cal_corr()
                    rmse_results_overall, rmse_results_hourly, rmse_results_monthly, \
                        rmse_results_overall_per, rmse_results_hourly_per, rmse_results_monthly_per = rmsecalc.calc_rmse()

                    mbe_results_overall, mbe_results_hourly, mbe_results_monthly, \
                        mbe_results_overall_per, mbe_results_hourly_per, mbe_results_monthly_per = mbecalc.calc_mbe()

                adap_table_tabletype = st.selectbox(
                    'Select the Data Type:',
                    ['Correlation', 'RMSE', 'MBE']
                )

                if adap_table_tabletype == 'Correlation':
                    st.header('Overall Correlation')
                    st.write(correlation_results_overall)

                    st.header('Monthly Correlation')
                    st.write(correlation_results_monthly)

                    st.header('Hourly Correlation')
                    st.write(correlation_results_hourly)

                if adap_table_tabletype == 'RMSE':
                    # with adap_tab_head1:
                    #     st.title('Overall RMSE Absolute and %')
                    st.header('Overall RMSE Absolute and %')
                    adap_tab_c1, adap_tab_c2 = st.columns(2)
                    with adap_tab_c1:
                        st.write(rmse_results_overall)
                    with adap_tab_c2:
                        st.write(rmse_results_overall_per)

                    st.header('Monthly RMSE Absolute and %')
                    adap_tab_c3, adap_tab_c4 = st.columns(2)
                    with adap_tab_c3:
                        st.write(rmse_results_monthly)
                    with adap_tab_c4:
                        st.write(rmse_results_monthly_per)

                    st.header('Hourly RMSE Absolute and %')
                    adap_tab_c5, adap_tab_c6 = st.columns(2)
                    with adap_tab_c5:
                        st.write(rmse_results_hourly)
                    with adap_tab_c6:
                        st.write(rmse_results_hourly_per)

                if adap_table_tabletype == 'MBE':
                    # with adap_tab_head1:
                    #     st.title('Overall RMSE Absolute and %')
                    st.header('Overall MBE Absolute and %')
                    adap_tab_c1, adap_tab_c2 = st.columns(2)
                    with adap_tab_c1:
                        st.write(mbe_results_overall)
                    with adap_tab_c2:
                        st.write(mbe_results_overall_per)

                    st.header('Monthly MBE Absolute and %')
                    adap_tab_c3, adap_tab_c4 = st.columns(2)
                    with adap_tab_c3:
                        st.write(mbe_results_monthly)
                    with adap_tab_c4:
                        st.write(mbe_results_monthly_per)

                    st.header('Hourly MBE Absolute and %')
                    adap_tab_c5, adap_tab_c6 = st.columns(2)
                    with adap_tab_c5:
                        st.write(mbe_results_hourly)
                    with adap_tab_c6:
                        st.write(mbe_results_hourly_per)

    with adap_graphtab:
        if adap_graphtab and st.session_state.calculation_button:
            combined_df_tmp = pd.merge(files_dict[Satellite_tab]["resampled_data"],
                                       files_dict[Measured_tab]["resampled_data"], on='Timestamp', how='inner')

            combined_df_tmp[mea_ghi_col] = pd.to_numeric(
                combined_df_tmp[mea_ghi_col], errors='coerce')
            combined_df_tmp[sat_ghi_col] = pd.to_numeric(
                combined_df_tmp[sat_ghi_col], errors='coerce')

            # Drop rows with NaN values in the GHI columns
            combined_df_tmp.dropna(
                subset=[mea_ghi_col, sat_ghi_col], inplace=True)
            combined_df_tmp.reset_index(inplace=True)

            # # Plot the graph
            fig_adap1 = px.line(combined_df_tmp, x='Timestamp',
                                y=[mea_ghi_col, sat_ghi_col])
            st.plotly_chart(fig_adap1)

            # Second_graph
            resampled_measured_data = files_dict[Measured_tab]['resampled_data']
            # st.write(resampled_measured_data)

            # Get the original measured data (1-minute) from the dataframe
            selected_measured_data = files_dict[Measured_tab]['df']
            sattelite_ghi_data = files_dict[Satellite_tab]['df']

            GHI_cols_tmp = [x for x in list(files_dict[Measured_tab]['df'].columns) if "GHI" in x ][0]
            DHI_cols_tmp = [x for x in list(files_dict[Measured_tab]['df'].columns) if "DHI" in x ][0]
            DNI_cols_tmp = [x for x in list(files_dict[Measured_tab]['df'].columns) if "DNI" in x ][0]
            Temp_cols_tmp = [x for x in list(files_dict[Measured_tab]['df'].columns) if "Temp" in x ][0]

            selected_measured_data.rename(columns={GHI_cols_tmp: 'GHIm', DNI_cols_tmp: 'DNIm',
                                                   DHI_cols_tmp: 'DHIm',  Temp_cols_tmp: 'Tempm'}, inplace=True)

            fig_adap2 = go.Figure()
            fig_adap2.add_trace(go.Scatter(x=resampled_measured_data.index,
                                           # Replace with your DNI column name
                                           # y=resampled_measured_data[mea_ghi_col]
                                           y=resampled_measured_data[mea_ghi_col],
                                           mode='lines',
                                           line=dict(color='orange'),
                                           name='Resampled Hourly GHI'))

            fig_adap2.add_trace(go.Scatter(x=resampled_measured_data.index,
                                           # Assuming GHIm column
                                           y=resampled_measured_data[DNI_cols_tmp],
                                           mode='lines',
                                           line=dict(color='light blue'),
                                           name='Resampled Hourly DNI (DNIm)'))

    # Add scatter trace for selected_measured_data GHIm (1-minute GHI)
            fig_adap2.add_trace(go.Scatter(x=selected_measured_data.index,
                                           # Assuming 1-min GHIm column
                                           y=selected_measured_data['GHIm'],
                                           mode='lines',
                                           line=dict(color='green'),
                                           name='1-Minute GHI'))

            # Set the title and labels
            fig_adap2.update_layout(
                title="GHI Over Time",
                xaxis_title="Time",
                yaxis_title="GHI",
            )
            #         # Plot the second figure (fig_adap2)
            st.plotly_chart(fig_adap2)

            # Graph_5
            fig_adap3 = go.Figure()

            fig_adap3.add_trace(go.Scatter(x=resampled_measured_data.index,
                                           # Replace with your DNI column name
                                           # y=resampled_measured_data[mea_ghi_col]
                                           y=resampled_measured_data[mea_ghi_col],
                                           mode='lines',
                                           line=dict(color='orange'),
                                           fill='tozeroy',
                                           name='Resampled Hourly GHI'))

            fig_adap3.add_trace(go.Scatter(x=selected_measured_data.index,
                                # Assuming 1-min GHIm column
                                           y=selected_measured_data['GHIm'],
                                           mode='lines',
                                           line=dict(color='green'),
                                           fill='tonexty',
                                           name='1-Minute GHI'))
            fig_adap3.update_layout(
                title="GHI Over Time",
                xaxis_title="Time",
                yaxis_title="GHI",
            )
            st.plotly_chart(fig_adap3)

            # Comparison of Satellite and Measured GHI (Common Duration)-graph
            fig_adap4 = go.Figure()

            adapted_ghi = common_df['ghi_adapted_1']
            adapted_ghi_approach = common_df['ghi_adapted_2']

            fig_adap4.add_trace(go.Scatter(x=files_dict[Satellite_tab]['df'].index,

                                           y=files_dict[Satellite_tab]['df']['GHI'],
                                           mode='lines',
                                           line=dict(color='blue'),
                                           name='Satellite GHI'))

            fig_adap4.add_trace(go.Scatter(x=selected_measured_data.index,
                                           # Replace with your DNI column name
                                           # y=files_dict[Satellite_tab]['df']['GHI']
                                           y=selected_measured_data['GHIm'],
                                           mode='lines',
                                           line=dict(color='purple'),
                                           name='1min-measured GHI'))

            fig_adap4.add_trace(go.Scatter(x=resampled_measured_data.index,

                                           y=resampled_measured_data[mea_ghi_col],
                                           mode='lines',
                                           line=dict(color='yellow'),
                                           name='hourly-resampled measured GHI'))

            fig_adap4.add_trace(go.Scatter(x=common_df.index,
                                           y=adapted_ghi,
                                           mode='lines',
                                           line=dict(color='yellow'),
                                           name='Adapted GHI'))

            fig_adap4.add_trace(go.Scatter(x=common_df.index,
                                           y=adapted_ghi_approach,
                                           mode='lines',
                                           line=dict(color='green'),
                                           name='Adapted GHI Approach 2'))

            fig_adap4.update_layout(title='Comparison of Satellite and Measured GHI (Common Duration)',
                                    xaxis_title="Time",
                                    yaxis_title="Global Horizantal Irradiance(GHI)",)

            st.plotly_chart(fig_adap4)
    with down_csvtab:
        sample_common_df = st.session_state.get('common_df', None)

        if sample_common_df is not None:
            st.write("The Adapted df is populated")
            # st.write(sample_common_df.index)
            # st.write(sample_common_df)

            # Provide options to select which adaptation method to use
            adaptation_method = st.selectbox(
                "Select Adaptation Method",
                options=["Adaptation Method 1", "Adaptation Method 2"],

            )

            # Prepare the corresponding column name based on the selected adaptation method
            if adaptation_method == "Adaptation Method 1":
                selected_adapted_col = [
                    col for col in sample_common_df.columns if '_adapted_1' in col]
            else:
                selected_adapted_col = [
                    col for col in sample_common_df.columns if '_adapted_2' in col]

            sample_common_df['Year'] = sample_common_df.index.year
            # sample_common_df['Year'] = sample_common_df.index.year
            sample_common_df['Month'] = sample_common_df.index.month
            sample_common_df['Day'] = sample_common_df.index.day
            sample_common_df['Hour'] = sample_common_df.index.hour
            sample_common_df['Minute'] = sample_common_df.index.minute

            selected_columns = ["Year", "Month",
                                "Day", "Hour", "Minute", "GHI", "DNI"]
            final_columns = selected_columns + selected_adapted_col

            # Create the final DataFrame for download
            df_to_download = sample_common_df[final_columns]

            # Prepare the Excel template with Meteo values and merge the selected columns
            measured_data_dict = None
            satellite_data_dict = None

            for key, value in files_dict.items():
                if value['data_type'] == "Measured":  # Check for measured data type
                    measured_data_dict = value['data_dict_df']
                    st.write()
                else:
                    satellite_data_dict = value['data_dict_df']
                    # data_type":"Satellite"

            if measured_data_dict is not None:
                print('The measured_data_dict is:', measured_data_dict)
                print('The satellite_data_dict is:', satellite_data_dict)
                # print('The measured_data_dict is:', satellite_data_dict)

                # Prepare Excel workbook
                current_dir = os.path.dirname(__file__)
                template_path = os.path.join(
                    current_dir, 'utilsfol', 'Meteo_Template', 'Meteo_template.xlsx')
                wb = load_workbook(template_path)
                ws = wb.active

                # Add Meteo summary values (latitude, longitude, etc.)
                site_name = measured_data_dict.loc[measured_data_dict['Key']
                                                   == 'Site name', 'Value'].values[0]
                time_zone = measured_data_dict.loc[measured_data_dict['Key']
                                                   == 'Time zone', 'Value'].values[0]
                latitude = measured_data_dict.loc[measured_data_dict['Key']
                                                  == 'Latitude', 'Value'].values[0]
                longitude = measured_data_dict.loc[measured_data_dict['Key']
                                                   == 'Longitude', 'Value'].values[0]
                summary_period_1 = satellite_data_dict.loc[satellite_data_dict['Key'] ==
                                                           'Data Starts Date', 'Value'].values[0].strftime('%Y-%m-%d %H:%M:%S')
                summary_period_2 = satellite_data_dict.loc[satellite_data_dict['Key'] ==
                                                           'Data Ends Date', 'Value'].values[0].strftime('%Y-%m-%d %H:%M:%S')
                summary_data = {
                    'Meteo hourly data': None,  # Placeholder for actual data
                    'Site': site_name,
                    'Country': 'Atacama, Chile',
                    # Example: assuming data type is present
                    'Data Source': 'Measured Adapted Data',
                    'Time step': 'Hour',
                    'Latitude': latitude,
                    'Longitude': longitude,
                    'Altitude': '1689',  # Example static value
                    'Time Zone': time_zone,
                    'Hour Shift': '0',
                    'Time Shift': '0',
                    'Summarization period': f"{summary_period_1} to {summary_period_2}"
                }
                # Write summary data to the worksheet starting from row 2
                for row_num, (key, value) in enumerate(summary_data.items(), start=1):
                    # Write key to column A
                    ws.cell(row=row_num, column=1, value=key)
                    # Write value to column B
                    ws.cell(row=row_num, column=2, value=value)

                # Write the selected data from common_df starting from row 15 (or any other row after the summary)
                # First, write the column names (header row)
                for col_num, column_name in enumerate(df_to_download.columns, start=1):
                    ws.cell(row=13, column=col_num, value=column_name)

                # Now write the actual data starting from row 13
                for row_num, row_data in enumerate(df_to_download.values, start=14):
                    for col_num, value in enumerate(row_data, start=1):
                        ws.cell(row=row_num, column=col_num, value=value)

                # Save the modified workbook into a BytesIO object for download
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                # Provide download button for the modified template with Meteo values and common_df data
                st.download_button(
                    label="Download Adapted Data",
                    data=output,
                    file_name="Adapted_Meteo_Data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        else:
            st.write(
                'No common_df available: Please click Apply Adaption before downloading')
