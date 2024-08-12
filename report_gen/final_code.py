#generating an Excel output with consolidated and formatted data.
import pandas as pd
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def process_csv_files(csv_path, skip_rows=10):
    # Using glob to find all CSV files in the directory
    csv_files = glob.glob(os.path.join(csv_path, "*.csv"))

    # Reading the first few rows from each CSV file and appending to temp_data
    temp_data = pd.DataFrame()
    for file_path in csv_files:
        data = pd.read_csv(file_path, skiprows=1, index_col=None, nrows=8, on_bad_lines='skip', encoding='latin-1')
        data = data.reset_index()
        temp_data = pd.concat([temp_data, data], ignore_index=True)

    # Rename columns in the temporary data
    temp_data.rename(columns={'index': 'project name', 'Unnamed: 0': 'File', 'File': 'File date', 
                              'File date': 'Description', 'Description': 'Empty'}, inplace=True)

    # Reading the remaining rows from each CSV file and storing in a list
    dfs = []
    for file_path in csv_files:
        df = pd.read_csv(file_path, header=[0, 1], skiprows=skip_rows, encoding='latin-1')
        dfs.append(df)

    # Concatenating all dataframes into a single dataframe
    final_df = pd.concat(dfs, ignore_index=True)

    # Resetting the index to generate a new sequential index
    final_df.reset_index(drop=True, inplace=True)

    # Update column names by combining column names with units
    columns_with_units = [
        f"{column.strip()} [{unit.strip()}]" for column, unit in zip(final_df.columns.get_level_values(0), final_df.columns.get_level_values(1))
    ]
    final_df.columns = columns_with_units

    # Changing the 'date []' column name to 'timestamp'
    final_df.rename(columns={'date []': 'timestamp'}, inplace=True)

    # Dropping the first row as it was used to create the new column names
    final_df.drop(index=0, inplace=True)

    # Resetting the index again
    final_df.reset_index(drop=True, inplace=True)

    # Convert 'timestamp' column to datetime format with handling invalid values
    final_df['timestamp'] = pd.to_datetime(final_df['timestamp'], format='%d/%m/%y %H:%M', errors='coerce')

    # Drop rows where 'timestamp' is NaT (invalid or empty values)
    final_df = final_df.dropna(subset=['timestamp'])

    # Extracting Year, Month, Day, and Hour from the 'timestamp' column
    final_df['Year'] = final_df['timestamp'].dt.year
    final_df['Month'] = final_df['timestamp'].dt.month
    final_df['Day'] = final_df['timestamp'].dt.day
    final_df['Hour'] = final_df['timestamp'].dt.hour

    # Formatting the 'timestamp' column as desired
    final_df['timestamp'] = final_df['timestamp'].dt.strftime('%d/%m/%Y %H:%M')

    # Reordering the columns
    new_order = ['timestamp'] + list(final_df.columns[-4:]) + list(final_df.columns[1:-4])
    final_df = final_df[new_order]

    return final_df

def match_columns_with_config(final_df, config_path):
    # Load the configuration file
    config_df = pd.read_excel(config_path)

    # Extract the list of columns and the configuration code list
    final_df_columns_list = final_df.columns.tolist()
    final_desc_list = []

    # Match columns with configuration codes
    for col in final_df_columns_list:
        if len(config_df[(config_df['Code'] == col)]) > 0:
            final_desc_list.append(config_df[config_df['Code'] == col]['Description'].to_list()[0])
        else:
            final_desc_list.append('')

    return final_desc_list, final_df_columns_list

def write_to_template(template_path, worksheet_name, start_row, final_df, final_desc_list):
    # Load the template
    wb = load_workbook(template_path)
    worksheet = wb[worksheet_name]

    # Write the header with descriptions
    for i, col_name in enumerate(final_desc_list):
        cell = worksheet.cell(row=start_row, column=i + 1, value=col_name)
        cell.fill = PatternFill(start_color="404F64", fill_type="solid")
        cell.font = Font(color="FFFFFF")

    # Write the actual column names
    for i, col_name in enumerate(final_df.columns):
        cell = worksheet.cell(row=start_row + 1, column=i + 1, value=col_name)
        cell.fill = PatternFill(start_color="404F64", fill_type="solid")
        cell.font = Font(color="FFFFFF")

    # Append the data rows
    for i, row in enumerate(final_df.values.tolist()):
        worksheet.append(row)
        cell_color = worksheet.cell(row=i + start_row + 2, column=1)
        cell_color.fill = PatternFill(start_color="B8CACC", fill_type="solid")

    # Save the workbook
    wb.save(template_path)
    print("Appended Successfully")

# Parameters for reuse
csv_path = r'path_to_csv_files'
template_path = r'path_to_new_template.xlsx'
worksheet_name = 'Sheet1'  # Update this to the correct sheet name in the new template
config_path = r'path_to_config_file.xlsx'
start_row = 6

# Execute the processing
final_df = process_csv_files(csv_path)
final_desc_list, final_df_columns_list = match_columns_with_config(final_df, config_path)
write_to_template(template_path, worksheet_name, start_row, final_df, final_desc_list)
